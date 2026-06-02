"""
maltego_to_xlsx.py
──────────────────
Reads a Maltego CSV export (mixed entity types) and produces a structured
Excel workbook with:
  - Sheet 1 : All Entities       (every row, normalised)
  - Sheet 2 : Social Accounts    (Twitter, Facebook, Instagram, LinkedIn …)
  - Sheet 3 : Email Addresses
  - Sheet 4 : Domains & URLs
  - Sheet 5 : People
  - Sheet 6 : Phone Numbers
  - Sheet 7 : Other Entities
  - Sheet 8 : Hygiene Flags      (one row per subject, 10 flag columns — fill manually)
  - Sheet 9 : Instructions

Usage:
  python maltego_to_xlsx.py <maltego_export.csv> [output.xlsx]
"""

import csv
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styling helpers ────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", start_color="0D1117")
SUB_FILL   = PatternFill("solid", start_color="0A1929")
FLAG_FILL  = PatternFill("solid", start_color="1A2535")
ALT_FILL   = PatternFill("solid", start_color="080B14")
HDR_FONT   = Font(name="Arial", bold=True, color="00D4FF", size=10)
SUB_FONT   = Font(name="Arial", bold=True, color="9DDDDD", size=10)
VAL_FONT   = Font(name="Arial", color="6A9090", size=9)
FLAG_FONT  = Font(name="Arial", color="FF6B8A", size=9)
OK_FONT    = Font(name="Arial", color="2A4060", size=9)
THIN       = Side(style="thin", color="1E293B")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = HDR_FONT
    c.fill      = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER
    return c

def data_cell(ws, row, col, value, bold=False, color="6A9090", fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=color, size=9)
    c.fill      = fill or (SUB_FILL if row % 2 == 0 else ALT_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = BORDER
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ── Entity type classification ─────────────────────────────────────────────
SOCIAL_TYPES = {
    "maltego.twitteraffiliate", "maltego.facebookobject", "maltego.instagramaffiliate",
    "maltego.linkedinaffiliate", "maltego.alias", "maltego.tiktokaffiliate",
    "maltego.youtubeaffiliate", "maltego.redditaffiliate", "maltego.githubaffiliate",
    "maltego.snapchataffiliate"
}
EMAIL_TYPES   = {"maltego.emailaddress"}
DOMAIN_TYPES  = {"maltego.domain", "maltego.website", "maltego.url", "maltego.netname"}
PERSON_TYPES  = {"maltego.person"}
PHONE_TYPES   = {"maltego.phonenumber"}

def classify(entity_type):
    t = entity_type.strip().lower()
    if t in SOCIAL_TYPES:  return "social"
    if t in EMAIL_TYPES:   return "email"
    if t in DOMAIN_TYPES:  return "domain"
    if t in PERSON_TYPES:  return "person"
    if t in PHONE_TYPES:   return "phone"
    return "other"

def infer_platform(row):
    """Guess the platform from entity type + network/affiliation fields."""
    t = row.get("Type", "").strip().lower()
    net = (row.get("maltego.affiliation.network", "") or
           row.get("affiliation.network", "") or "").strip()
    if net:
        return net
    mapping = {
        "maltego.twitteraffiliate":   "Twitter",
        "maltego.facebookobject":     "Facebook",
        "maltego.instagramaffiliate": "Instagram",
        "maltego.linkedinaffiliate":  "LinkedIn",
        "maltego.tiktokaffiliate":    "TikTok",
        "maltego.youtubeaffiliate":   "YouTube",
        "maltego.redditaffiliate":    "Reddit",
        "maltego.githubaffiliate":    "GitHub",
        "maltego.snapchataffiliate":  "Snapchat",
        "maltego.alias":              "Unknown",
    }
    return mapping.get(t, "")

def get_value(row):
    """Pull the primary display value from a Maltego row."""
    for key in ["Value", "value", "Entity Value", "entity.value"]:
        if row.get(key, "").strip():
            return row[key].strip()
    return ""

def get_name(row):
    for key in ["entity.name", "person.firstname", "person.fullname"]:
        if row.get(key, "").strip():
            v = row[key].strip()
            last = row.get("person.lastname", "").strip()
            if last and key == "person.firstname":
                return f"{v} {last}"
            return v
    return get_value(row)

# ── Parse Maltego CSV ──────────────────────────────────────────────────────
def parse_maltego_csv(path):
    records = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            entity_type = row.get("Type", row.get("type", "")).strip()
            if not entity_type:
                continue
            records.append({
                "type":       entity_type,
                "category":   classify(entity_type),
                "value":      get_value(row),
                "name":       get_name(row),
                "platform":   infer_platform(row),
                "email":      row.get("email.address", "").strip(),
                "screen":     row.get("twitter.screenname", row.get("maltego.affiliation.uid", "")).strip(),
                "domain":     row.get("fqdn", "").strip(),
                "url":        row.get("url", "").strip(),
                "phone":      row.get("phonenum", "").strip(),
                "weight":     row.get("weight", "").strip(),
                "link_label": row.get("Link Label", row.get("link.label", "")).strip(),
                "_raw":       dict(row),
            })
    return records

# ── Write Sheets ───────────────────────────────────────────────────────────
def write_all_entities(wb, records):
    ws = wb.active
    ws.title = "All Entities"
    headers = ["#", "Entity Type", "Category", "Value", "Name / Handle",
               "Platform", "Email", "Domain / URL", "Phone", "Link Label"]
    for c, h in enumerate(headers, 1):
        hdr_cell(ws, 1, c, h)
    ws.row_dimensions[1].height = 28

    for i, r in enumerate(records, 1):
        row = i + 1
        data_cell(ws, row, 1, i)
        data_cell(ws, row, 2, r["type"],       color="4A7090")
        data_cell(ws, row, 3, r["category"].upper(), color="3A6080")
        data_cell(ws, row, 4, r["value"],      color="9DDDDD", bold=True)
        data_cell(ws, row, 5, r["name"])
        data_cell(ws, row, 6, r["platform"],   color="6AA090")
        data_cell(ws, row, 7, r["email"])
        data_cell(ws, row, 8, r["domain"] or r["url"])
        data_cell(ws, row, 9, r["phone"])
        data_cell(ws, row, 10, r["link_label"], color="3A5060")

    set_col_widths(ws, [5, 28, 12, 30, 24, 16, 28, 34, 16, 20])
    freeze(ws)

def write_filtered_sheet(wb, title, records, category, headers, row_fn):
    ws = wb.create_sheet(title)
    for c, h in enumerate(headers, 1):
        hdr_cell(ws, 1, c, h)
    ws.row_dimensions[1].height = 28
    subset = [r for r in records if r["category"] == category]
    for i, r in enumerate(subset, 1):
        row_fn(ws, i + 1, r)
    freeze(ws)
    return ws

def write_social(wb, records):
    headers = ["#", "Handle / Value", "Platform", "Entity Type", "Link Label"]
    def row_fn(ws, row, r):
        data_cell(ws, row, 1, row - 1)
        data_cell(ws, row, 2, r["screen"] or r["value"], color="9DDDDD", bold=True)
        data_cell(ws, row, 3, r["platform"], color="6AA090")
        data_cell(ws, row, 4, r["type"],     color="3A6080")
        data_cell(ws, row, 5, r["link_label"])
    ws = write_filtered_sheet(wb, "Social Accounts", records, "social", headers, row_fn)
    set_col_widths(ws, [5, 28, 16, 30, 20])

def write_emails(wb, records):
    headers = ["#", "Email Address", "Provider", "Link Label"]
    def row_fn(ws, row, r):
        addr = r["email"] or r["value"]
        provider = addr.split("@")[-1] if "@" in addr else ""
        data_cell(ws, row, 1, row - 1)
        data_cell(ws, row, 2, addr,      color="9DDDDD", bold=True)
        data_cell(ws, row, 3, provider,  color="6AA090")
        data_cell(ws, row, 4, r["link_label"])
    ws = write_filtered_sheet(wb, "Email Addresses", records, "email", headers, row_fn)
    set_col_widths(ws, [5, 34, 20, 20])

def write_domains(wb, records):
    headers = ["#", "Domain / URL", "Entity Type", "Link Label"]
    def row_fn(ws, row, r):
        data_cell(ws, row, 1, row - 1)
        data_cell(ws, row, 2, r["url"] or r["domain"] or r["value"], color="9DDDDD", bold=True)
        data_cell(ws, row, 3, r["type"],      color="3A6080")
        data_cell(ws, row, 4, r["link_label"])
    ws = write_filtered_sheet(wb, "Domains & URLs", records, "domain", headers, row_fn)
    set_col_widths(ws, [5, 44, 26, 20])

def write_people(wb, records):
    headers = ["#", "Full Name", "Entity Type", "Link Label"]
    def row_fn(ws, row, r):
        data_cell(ws, row, 1, row - 1)
        data_cell(ws, row, 2, r["name"] or r["value"], color="9DDDDD", bold=True)
        data_cell(ws, row, 3, r["type"],   color="3A6080")
        data_cell(ws, row, 4, r["link_label"])
    ws = write_filtered_sheet(wb, "People", records, "person", headers, row_fn)
    set_col_widths(ws, [5, 28, 26, 20])

def write_phones(wb, records):
    headers = ["#", "Phone Number", "Entity Type", "Link Label"]
    def row_fn(ws, row, r):
        data_cell(ws, row, 1, row - 1)
        data_cell(ws, row, 2, r["phone"] or r["value"], color="9DDDDD", bold=True)
        data_cell(ws, row, 3, r["type"],   color="3A6080")
        data_cell(ws, row, 4, r["link_label"])
    ws = write_filtered_sheet(wb, "Phone Numbers", records, "phone", headers, row_fn)
    set_col_widths(ws, [5, 22, 26, 20])

def write_other(wb, records):
    headers = ["#", "Value", "Entity Type", "Link Label"]
    def row_fn(ws, row, r):
        data_cell(ws, row, 1, row - 1)
        data_cell(ws, row, 2, r["value"], color="9DDDDD", bold=True)
        data_cell(ws, row, 3, r["type"],  color="3A6080")
        data_cell(ws, row, 4, r["link_label"])
    ws = write_filtered_sheet(wb, "Other Entities", records, "other", headers, row_fn)
    set_col_widths(ws, [5, 34, 30, 20])

def write_hygiene_flags(wb, records):
    """Blank flag sheet — one row per unique social/email subject for manual fill."""
    ws = wb.create_sheet("Hygiene Flags")
    headers = [
        "Subject (Value)", "Entity Type", "Platform",
        "Password_Reuse", "MFA_Disabled", "Email_Exposed",
        "Data_Breach", "Weak_Username", "Over_Sharing",
        "Old_Accounts", "HTTP_Domains", "Geolocation_Exposed",
        "Phishing_Susceptible"
    ]
    for c, h in enumerate(headers, 1):
        hdr_cell(ws, 1, c, h)
    ws.row_dimensions[1].height = 32

    seen = set()
    row = 2
    for r in records:
        if r["category"] not in ("social", "email"):
            continue
        key = r["value"]
        if key in seen:
            continue
        seen.add(key)

        data_cell(ws, row, 1, r["value"],    color="9DDDDD", bold=True, fill=SUB_FILL)
        data_cell(ws, row, 2, r["type"],     color="3A6080", fill=SUB_FILL)
        data_cell(ws, row, 3, r["platform"], color="6AA090", fill=SUB_FILL)
        for col in range(4, 14):
            c = ws.cell(row=row, column=col, value="")
            c.fill   = FLAG_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    note_row = row + 1
    nc = ws.cell(row=note_row, column=1,
                 value="Enter 1 = indicator present,  0 = not present,  leave blank = unknown")
    nc.font = Font(name="Arial", italic=True, color="2A4060", size=9)

    set_col_widths(ws, [30, 28, 16, 16, 14, 15, 13, 15, 14, 13, 13, 20, 20])
    freeze(ws)

def write_instructions(wb):
    ws = wb.create_sheet("Instructions")

    lines = [
        ("MALTEGO EVIDENCE → EXCEL PROCESSOR", None, "title"),
        ("", None, "blank"),
        ("WORKFLOW", None, "section"),
        ("1. Run your Maltego transforms to build the graph.", None, "body"),
        ("2. Export the graph: File → Export → Export Graph as CSV.", None, "body"),
        ("3. Run:  python maltego_to_xlsx.py your_export.csv output.xlsx", None, "code"),
        ("4. Open the generated Excel file.", None, "body"),
        ("5. Review each typed sheet (Social, Email, Domains, People, Phones, Other).", None, "body"),
        ("6. Go to 'Hygiene Flags' and fill in 1/0 for each indicator per subject.", None, "body"),
        ("", None, "blank"),
        ("SHEETS", None, "section"),
        ("All Entities",     "Every entity from the CSV, normalised into a single table.", "item"),
        ("Social Accounts",  "Twitter, Facebook, Instagram, LinkedIn, GitHub, etc.", "item"),
        ("Email Addresses",  "All maltego.EmailAddress entities.", "item"),
        ("Domains & URLs",   "Domains, Websites, URLs.", "item"),
        ("People",           "maltego.Person entities.", "item"),
        ("Phone Numbers",    "maltego.PhoneNumber entities.", "item"),
        ("Other Entities",   "Any entity type not covered above.", "item"),
        ("Hygiene Flags",    "Manual flag sheet — one row per social/email subject.", "item"),
        ("", None, "blank"),
        ("HYGIENE FLAG DEFINITIONS", None, "section"),
        ("Password_Reuse",       "Subject reuses passwords across platforms", "flag"),
        ("MFA_Disabled",         "Multi-factor authentication not enabled", "flag"),
        ("Email_Exposed",        "Email address visible publicly on profile", "flag"),
        ("Data_Breach",          "Account found in known data breach databases", "flag"),
        ("Weak_Username",        "Predictable or guessable username pattern", "flag"),
        ("Over_Sharing",         "Excessive personal info shared publicly", "flag"),
        ("Old_Accounts",         "Dormant or inactive accounts still active", "flag"),
        ("HTTP_Domains",         "Linked to non-HTTPS (insecure) domains", "flag"),
        ("Geolocation_Exposed",  "Location metadata exposed in posts or profile", "flag"),
        ("Phishing_Susceptible", "Linked to phishing indicators or spoofed pages", "flag"),
        ("", None, "blank"),
        ("SUPPORTED MALTEGO ENTITY TYPES", None, "section"),
        ("Social",  "TwitterAffiliate, FacebookObject, InstagramAffiliate, LinkedInAffiliate, Alias, TikTokAffiliate, YouTubeAffiliate, RedditAffiliate, GitHubAffiliate, SnapchatAffiliate", "item"),
        ("Email",   "EmailAddress", "item"),
        ("Domain",  "Domain, Website, URL, Netname", "item"),
        ("Person",  "Person", "item"),
        ("Phone",   "PhoneNumber", "item"),
        ("Other",   "Any unrecognised entity type is captured in the Other sheet", "item"),
    ]

    for r_idx, (label, desc, style) in enumerate(lines, 1):
        lc = ws.cell(row=r_idx, column=1, value=label)
        if style == "title":
            lc.font = Font(name="Arial", bold=True, color="00D4FF", size=12)
            lc.fill = HDR_FILL
        elif style == "section":
            lc.font = Font(name="Arial", bold=True, color="00D4FF", size=10)
            lc.fill = PatternFill("solid", start_color="0A1929")
        elif style == "code":
            lc.font = Font(name="Courier New", color="FFD60A", size=9)
            lc.fill = PatternFill("solid", start_color="0A0F00")
        elif style == "flag":
            lc.font = Font(name="Arial", bold=True, color="9DDDDD", size=9)
            lc.fill = SUB_FILL
        elif style == "item":
            lc.font = Font(name="Arial", bold=True, color="4A7090", size=9)
            lc.fill = ALT_FILL
        else:
            lc.font = Font(name="Arial", color="6A9090", size=9)
            lc.fill = ALT_FILL

        if desc:
            dc = ws.cell(row=r_idx, column=2, value=desc)
            dc.font = Font(name="Arial", color="4A7090", size=9)
            dc.fill = ALT_FILL

        ws.row_dimensions[r_idx].height = 18

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70

# ── Main ───────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python maltego_to_xlsx.py <input.csv> [output.xlsx]")
        sys.exit(1)

    in_path  = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(in_path)[0] + "_processed.xlsx"

    print(f"Reading: {in_path}")
    records = parse_maltego_csv(in_path)
    print(f"  → {len(records)} entities found")

    wb = Workbook()
    write_all_entities(wb, records)
    write_social(wb, records)
    write_emails(wb, records)
    write_domains(wb, records)
    write_people(wb, records)
    write_phones(wb, records)
    write_other(wb, records)
    write_hygiene_flags(wb, records)
    write_instructions(wb)

    wb.save(out_path)
    print(f"Saved:   {out_path}")

    cats = {}
    for r in records:
        cats[r["category"]] = cats.get(r["category"], 0) + 1
    print("\nBreakdown:")
    for k, v in sorted(cats.items()):
        print(f"  {k:<10} {v}")

if __name__ == "__main__":
    main()
